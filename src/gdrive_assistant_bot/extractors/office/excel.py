from __future__ import annotations

import io
from typing import Any

import xlrd
from openpyxl import load_workbook

from ..base import ExtractedContent, ExtractionContext, FileExtractor

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLS_MIME_TYPE = "application/vnd.ms-excel"


class XlsxExtractor(FileExtractor):
    """Extract text from XLSX files."""

    @property
    def mime_types(self) -> list[str]:
        return [XLSX_MIME_TYPE]

    @property
    def file_extensions(self) -> list[str]:
        return ["xlsx"]

    def can_extract(self, file_meta: dict[str, Any]) -> bool:
        if file_meta.get("mimeType") == XLSX_MIME_TYPE:
            return True
        return self._extension(file_meta) == "xlsx"

    def extract(self, file_meta: dict[str, Any], context: ExtractionContext) -> ExtractedContent:
        size = self._to_int(file_meta.get("size"))
        max_bytes = int(context.settings.OFFICE_MAX_FILE_SIZE_MB * 1024 * 1024)
        if size and size > max_bytes:
            return ExtractedContent(
                text="", file_type="xlsx", metadata={"skipped": "size_limit", "size_bytes": size}
            )

        xlsx_bytes = context.download_binary(file_meta["id"])
        text = self._extract_xlsx(
            xlsx_bytes,
            max_sheets=context.settings.EXCEL_MAX_SHEETS,
            max_rows=context.settings.STORAGE_GOOGLE_DRIVE_MAX_ROWS_PER_SHEET,
        )
        return ExtractedContent(
            text=text.strip(),
            file_type="xlsx",
            metadata={"mime_type": file_meta.get("mimeType"), "file_size_bytes": len(xlsx_bytes)},
        )

    @staticmethod
    def _extract_xlsx(xlsx_bytes: bytes, *, max_sheets: int, max_rows: int) -> str:
        workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet_index, sheet_name in enumerate(workbook.sheetnames, start=1):
            if max_sheets and sheet_index > max_sheets:
                lines.append(f"... (limited to {max_sheets} sheets)")
                break

            sheet = workbook[sheet_name]
            lines.append(f"=== SHEET: {sheet_name} ===")
            for row_index, row in enumerate(
                sheet.iter_rows(max_row=max_rows, values_only=True), start=1
            ):
                values = [
                    str(value).strip() for value in row if value is not None and str(value).strip()
                ]
                if values:
                    lines.append("\t".join(values))
                if max_rows and row_index >= max_rows:
                    break
            if max_rows and sheet.max_row and sheet.max_row > max_rows:
                lines.append(f"... (limited to {max_rows} rows, {sheet.max_row} total)")
            lines.append("")
        return "\n".join(lines)

    @staticmethod
    def _extension(file_meta: dict[str, Any]) -> str | None:
        ext = file_meta.get("fileExtension")
        if isinstance(ext, str) and ext.strip():
            return ext.lower().lstrip(".")

        name = file_meta.get("name")
        if not isinstance(name, str) or "." not in name:
            return None
        return name.rsplit(".", 1)[-1].strip().lower() or None

    @staticmethod
    def _to_int(value: Any) -> int | None:
        if isinstance(value, int):
            return value
        if isinstance(value, str) and value.isdigit():
            return int(value)
        return None


class XlsExtractor(FileExtractor):
    """Extract text from legacy XLS files."""

    @property
    def mime_types(self) -> list[str]:
        return [_XLS_MIME_TYPE]

    @property
    def file_extensions(self) -> list[str]:
        return ["xls"]

    def can_extract(self, file_meta: dict[str, Any]) -> bool:
        if file_meta.get("mimeType") == _XLS_MIME_TYPE:
            return True
        return self._extension(file_meta) == "xls"

    def extract(self, file_meta: dict[str, Any], context: ExtractionContext) -> ExtractedContent:
        size = self._to_int(file_meta.get("size"))
        max_bytes = int(context.settings.OFFICE_MAX_FILE_SIZE_MB * 1024 * 1024)
        if size and size > max_bytes:
            return ExtractedContent(
                text="", file_type="xls", metadata={"skipped": "size_limit", "size_bytes": size}
            )

        xls_bytes = context.download_binary(file_meta["id"])
        text = self._extract_xls(
            xls_bytes,
            max_sheets=context.settings.EXCEL_MAX_SHEETS,
            max_rows=context.settings.STORAGE_GOOGLE_DRIVE_MAX_ROWS_PER_SHEET,
        )
        return ExtractedContent(
            text=text.strip(),
            file_type="xls",
            metadata={"mime_type": file_meta.get("mimeType"), "file_size_bytes": len(xls_bytes)},
        )

    @staticmethod
    def _extract_xls(xls_bytes: bytes, *, max_sheets: int, max_rows: int) -> str:
        workbook = xlrd.open_workbook(file_contents=xls_bytes)
        lines: list[str] = []
        for sheet_index in range(workbook.nsheets):
            if max_sheets and sheet_index >= max_sheets:
                lines.append(f"... (limited to {max_sheets} sheets)")
                break

            sheet = workbook.sheet_by_index(sheet_index)
            lines.append(f"=== SHEET: {sheet.name} ===")
            row_limit = min(sheet.nrows, max_rows) if max_rows else sheet.nrows
            for row_idx in range(row_limit):
                values: list[str] = []
                for col_idx in range(sheet.ncols):
                    value = sheet.cell_value(row_idx, col_idx)
                    cell = str(value).strip()
                    if cell:
                        values.append(cell)
                if values:
                    lines.append("\t".join(values))
            if max_rows and sheet.nrows > max_rows:
                lines.append(f"... (limited to {max_rows} rows, {sheet.nrows} total)")
            lines.append("")
        return "\n".join(lines)

    @staticmethod
    def _extension(file_meta: dict[str, Any]) -> str | None:
        ext = file_meta.get("fileExtension")
        if isinstance(ext, str) and ext.strip():
            return ext.lower().lstrip(".")

        name = file_meta.get("name")
        if not isinstance(name, str) or "." not in name:
            return None
        return name.rsplit(".", 1)[-1].strip().lower() or None

    @staticmethod
    def _to_int(value: Any) -> int | None:
        if isinstance(value, int):
            return value
        if isinstance(value, str) and value.isdigit():
            return int(value)
        return None
